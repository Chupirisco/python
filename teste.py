import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
import time
from datetime import datetime

# Lista de estados brasileiros
estados_brasileiros = ['AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO']

def calcular_idade(dia, mes, ano):
    hoje = datetime.today()
    data_formatada = datetime(int(ano), int(mes), int(dia))
    idade = hoje.year - data_formatada.year - ((hoje.month, hoje.day) < (data_formatada.month, data_formatada.day))
    return idade

def cadastrar_aluno(dados_aluno):
    try:
        # Abre o arquivo existente ou cria um novo se não existir
        wb = Workbook()
        try:
            wb = load_workbook("matriculas_alunos.xlsx")
        except FileNotFoundError:
            pass

        # Seleciona a planilha ativa
        ws = wb.active

        # Adiciona os dados do aluno
        ws.append(dados_aluno)

        # Salva a planilha
        wb.save("matriculas_alunos.xlsx")

        # Exibe mensagem de status
        sg.popup('Matrícula cadastrada com sucesso!')
    except Exception as e:
        sg.popup_error(f"Erro ao cadastrar a matrícula: {e}")

def excluir_matricula():
    try:
        wb = load_workbook("matriculas_alunos.xlsx")
        ws = wb.active
        ws.delete_rows(ws.max_row)
        wb.save("matriculas_alunos.xlsx")
        sg.popup('Matrícula excluída com sucesso!')
    except Exception as e:
        sg.popup_error(f"Erro ao excluir a matrícula: {e}")

sg.theme('DarkBlue')  # Define o tema para DarkGreen5

# Layout para os dados do aluno
layout_aluno = [
    [sg.Text('DADOS PESSOAIS', font=('Helvetica', 16), text_color='white')],
    [sg.Text('Nome', size=(15, 1)), sg.Input(key='nome', pad=((10, 10), (0, 0)))],
    [sg.Text('Data de Nascimento', size=(15, 1)), sg.Input(key='dia_nascimento', size=(3, 1)), sg.Text('/', size=(1, 1)), sg.Input(key='mes_nascimento', size=(3, 1)), sg.Text('/', size=(1, 1)), sg.Input(key='ano_nascimento', size=(5, 1))],
    [sg.Text('Cidade Natal', size=(15, 1)), sg.Input(key='cidade_natal', pad=((10, 10), (0, 0)))],
    [sg.Text('UF Natal', size=(15, 1)), sg.Combo(estados_brasileiros, key='uf_natal', pad=((10, 10), (0, 0)))],
    [sg.Text('Sexo', size=(15, 1)), sg.Radio('Masculino', "SEXO", key='sexo_m', default=True), sg.Radio('Feminino', "SEXO", key='sexo_f')],
    [sg.Text('Nome do Pai', size=(15, 1)), sg.Input(key='nome_pai', pad=((10, 10), (0, 0)))],
    [sg.Text('Nome da Mãe', size=(15, 1)), sg.Input(key='nome_mae', pad=((10, 10), (0, 0)))],
    [sg.Text('CPF', size=(15, 1)), sg.Input(key='cpf', pad=((10, 10), (0, 0)))],
    [sg.Text('SUS', size=(15, 1)), sg.Input(key='sus', pad=((10, 10), (0, 0)))],
    [sg.Text('ID', size=(15, 1)), sg.Input(key='id', pad=((10, 10), (0, 0)))],
    [sg.Text('Telefone', size=(15, 1)), sg.Input(key='telefone', pad=((10, 10), (0, 0)))],
    [sg.Text('ENDEREÇO', font=('Helvetica', 16), text_color='white')],
    [sg.Text('Rua', size=(15, 1)), sg.Input(key='rua', pad=((10, 10), (0, 0)))],
    [sg.Text('Número', size=(15, 1)), sg.Input(key='numero', pad=((10, 10), (0, 0)))],
    [sg.Text('Bairro', size=(15, 1)), sg.Input(key='bairro', pad=((10, 10), (0, 0)))],
    [sg.Text('Cidade', size=(15, 1)), sg.Input(key='cidade_endereco', pad=((10, 10), (0, 0)))],
    [sg.Text('UF', size=(15, 1)), sg.Combo(estados_brasileiros, key='uf_endereco', pad=((10, 10), (0, 0)))],
    [sg.Text('MATRÍCULA', font=('Helvetica', 16), text_color='white')],
    [sg.Text('Curso', size=(15, 1)), sg.Input(key='curso', pad=((10, 10), (0, 0)))],
    [sg.Text('Data de Matrícula', size=(15, 1)), sg.Input(key='dia_matricula', size=(3, 1)), sg.Text('/', size=(1, 1)), sg.Input(key='mes_matricula', size=(3, 1)), sg.Text('/', size=(1, 1)), sg.Input(key='ano_matricula', size=(5, 1))],
    [sg.Text('Série/Ano', size=(15, 1)), sg.Input(key='serie_ano', pad=((10, 10), (0, 0)))],
    [sg.Text('Turno', size=(15, 1)), sg.Combo(['Matutino', 'Vespertino'], key='turno', default_value='Matutino', pad=((10, 10), (0, 0)))],
    [sg.Text('Situação', size=(15, 1)), sg.Combo(['Matriculado', 'Transferido'], key='situacao', default_value='Matriculado', pad=((10, 10), (0, 0)))],
    [sg.Button('Cadastrar', button_color=('white', 'blue')), sg.Button('Nova Matrícula', button_color=('white', 'blue')), sg.Button('Excluir Matrícula', button_color=('white', 'red'))],
    [sg.Text(size=(40, 1),

 key='-STATUS-', text_color='white')],  # Elemento para exibir a mensagem de status
]

#classe para verificar se o CPF é válido
class Verificacao:
    def VerificaCpf(self, cpf):
        if len(cpf) != 11:
            return False

        # Verificando se todos os dígitos são iguais
        if cpf == cpf[0] * 11:
            return False

        # Calculando o primeiro dígito verificador
        soma = 0
        for i in range(9):
            soma += int(cpf[i]) * (10 - i)
        resto = soma % 11
        if resto < 2:
            digito_verificador1 = 0
        else:
            digito_verificador1 = 11 - resto

        # Verificando o primeiro dígito verificador
        if digito_verificador1 != int(cpf[9]):
            return False

        # Calculando o segundo dígito verificador
        soma = 0
        for i in range(10):
            soma += int(cpf[i]) * (11 - i)
        resto = soma % 11
        if resto < 2:
            digito_verificador2 = 0
        else:
            digito_verificador2 = 11 - resto

        # Verificando o segundo dígito verificador
        if digito_verificador2 != int(cpf[10]):
            return False

        return True

# Criando uma instância da classe MyApp
minha_app = Verificacao()
# Criar a janela para os dados do aluno
janela_aluno = sg.Window('Cadastro de Aluno', layout_aluno, size=(900, 800))

# Loop de eventos para a tela de cadastro do aluno

while True:
    eventos_aluno, valores_aluno = janela_aluno.read()
    if eventos_aluno == sg.WINDOW_CLOSED:
        break
    if eventos_aluno == 'Cadastrar':
        #verifica c o cpf existe
        teste = minha_app.VerificaCpf(valores_aluno['cpf'])

        if teste == True:
            cpf = valores_aluno['cpf']        
            sexo = 'Masculino' if valores_aluno['sexo_m'] else 'Feminino'
            dados_aluno = [
                valores_aluno['nome'],
                f"{valores_aluno['dia_nascimento']}/{valores_aluno['mes_nascimento']}/{valores_aluno['ano_nascimento']}",
                calcular_idade(valores_aluno['dia_nascimento'], valores_aluno['mes_nascimento'], valores_aluno['ano_nascimento']),
                valores_aluno['cidade_natal'],
                valores_aluno['uf_natal'],
                sexo,
                valores_aluno['nome_pai'],
                valores_aluno['nome_mae'],
                valores_aluno['cpf'],
                valores_aluno['sus'],
                valores_aluno['id'],
                valores_aluno['telefone'],
                valores_aluno['rua'],
                valores_aluno['numero'],
                valores_aluno['bairro'],
                valores_aluno['cidade_endereco'],
                valores_aluno['uf_endereco'],
                valores_aluno['curso'],
                f"{valores_aluno['dia_matricula']}/{valores_aluno['mes_matricula']}/{valores_aluno['ano_matricula']}",
                valores_aluno['serie_ano'],
                valores_aluno['turno'],
                valores_aluno['situacao'],      
            ]
            cadastrar_aluno(dados_aluno)
            janela_aluno['-STATUS-'].update('Matrícula cadastrada com sucesso!')
        elif teste == False:
            sg.popup("Cpf Invalido")
        
    if eventos_aluno == 'Nova Matrícula':
        janela_aluno['Nova Matrícula'].update(button_color=('black', 'blue'))
        time.sleep(0.1)
        janela_aluno['Nova Matrícula'].update(button_color=('white', 'blue'))
        for key in valores_aluno.keys():
            janela_aluno[key]('')
    if eventos_aluno == 'Excluir Matrícula':
        janela_aluno['Excluir Matrícula'].update(button_color=('black', 'red'))
        time.sleep(0.1)
        janela_aluno['Excluir Matrícula'].update(button_color=('white', 'red'))
        excluir_matricula()

janela_aluno.close()
